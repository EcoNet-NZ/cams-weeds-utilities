#!/usr/bin/env python3
"""
Weed Visits Analyzer

Analyzes WeedLocations feature layer and its relationship with Visits_Table.
Compares multiple fields between WeedLocations and the latest visit record
to identify data synchronization issues using metadata-driven field comparison rules.

Key Features:
- Metadata-driven comparison rules for easy addition of new field checks
- Configurable reference date logic for determining "latest" visit
- VisitDataSource tracking to identify source of mismatches
- Detailed mismatch explanations with actual field values
- Supports custom ignore conditions for special cases (e.g., Purple statuses, bulk load dates)
- **Data Correction Mode**: Automatically fix mismatches and data quality issues

Three Rule Types:

1. FIELD_COMPARISON_RULES - Compare WeedLocations ↔ Visits_Table
   Used to identify when WeedLocations fields don't match their latest visit data

2. VISIT_CORRECTION_RULES - Correct Visits_Table internal fields (latest visits only)
   Used to fix data quality issues within Visits_Table itself
   Example: Set DateCheck from CreationDate_1 when DateCheck is missing
   Applied only to the LATEST visit per weed location

3. VISIT_FROM_WEED_RULES - Copy from WeedLocations to latest Visit
   Used to update the LATEST visit's fields based on WeedLocations data
   Example: Copy ParentStatusWithDomain → WeedVisitStatus when WeedVisitStatus is empty
   Applied only to the latest visit per weed location

Rule Metadata Structure (FIELD_COMPARISON_RULES):
- weed_field: Field name in WeedLocations layer
- visit_field: Field name in Visits_Table
- display_name: Human-readable name for reporting
- mismatch_column: Column name for mismatch indicator
- ignore_condition: Optional lambda(weed_val, visit_val) for custom skip logic
- description: Explanation of what the rule checks
- category: Optional category (e.g., 'audit') for grouping/filtering

Rule Metadata Structure (VISIT_CORRECTION_RULES):
- target_field: Field in Visits_Table to update
- source_field: Field in Visits_Table to copy value from
- display_name: Human-readable name for reporting
- condition: Lambda(row) that returns True if correction should be applied
- description: Explanation of what the rule corrects

Rule Metadata Structure (VISIT_FROM_WEED_RULES):
- visit_field: Field in Visits_Table (latest visit) to update
- weed_field: Field in WeedLocations to copy value from
- display_name: Human-readable name for reporting
- condition: Lambda(row) that returns True if correction should be applied (row has both weed and visit fields)
- description: Explanation of what the rule corrects

How "Latest" Visit is Determined:
  The script determines the SINGLE latest visit per weed location using:
  
  REFERENCE_DATE_STRATEGY = 'DateCheck → CreationDate_1'
  
  Logic:
  1. Load ALL visits for ALL weed locations from Visits_Table
  2. Group visits by GUID_visits (linking to WeedLocation's GlobalID)
  3. For each weed location:
     a. If visits with DateCheck exist → pick visit with most recent DateCheck
     b. Otherwise, if visits with CreationDate_1 exist → pick visit with most recent CreationDate_1
     c. Otherwise → no visit (null)
  4. ALL field comparisons and corrections use this SAME "latest" visit
  
  Result: Each WeedLocation compares against exactly ONE visit (the latest).
  
Ignore Conditions Apply at Field Level:
  When an ignore_condition is triggered for a field, it ONLY affects that specific field.
  Other rules continue to process the same row independently.
  
  Example: If Status has ignore_condition for Purple values:
  - Status field: Purple → Active (IGNORED, no mismatch flagged)
  - Urgency field: High → Medium (NOT ignored, mismatch flagged & correctable)
  - Same row, different outcomes per field

Correction Modes:

1. WeedLocations from Visits (--correct-weed-from-visits):
   Updates WeedLocations fields to match corresponding Visit data from the latest visit.
   Uses FIELD_COMPARISON_RULES to determine what to update.
   
2. Visits Internal (--correct-visits):
   Fixes data quality issues within Visits_Table using VISIT_CORRECTION_RULES.
   Example: Sets DateCheck from CreationDate_1 for visits missing DateCheck.
   Applied only to the LATEST visit per weed location.
   
3. Visits from WeedLocations (--correct-visits-from-weed):
   Updates latest Visit fields based on WeedLocations data using VISIT_FROM_WEED_RULES.
   Example: Copies ParentStatusWithDomain → WeedVisitStatus when WeedVisitStatus is empty.
   Applied only to the LATEST visit per weed location.

4. All Corrections (--correct-all):
   Applies all three correction types in optimal order:
   a. Visits internal corrections (--correct-visits)
   b. Visits from WeedLocations corrections (--correct-visits-from-weed)
   c. WeedLocations from Visits corrections (--correct-weed-from-visits)

Safety features include:
- Preview mode (--preview): Shows what would change without making changes
- Field filtering (--fields): Correct only specific WeedLocations fields
- Confirmation prompt: Requires explicit 'YES' before applying changes
- Detailed logging: Saves correction log with before/after values and update status
- Respects ignore conditions: Won't correct fields with special ignore logic (e.g., bulk load dates)

Usage Examples:
  # List all available rules
  python weed_visits_analyzer.py --list-fields
  
  # Preview all corrections
  python weed_visits_analyzer.py --env development --correct-all --preview
  
  # Apply all corrections
  python weed_visits_analyzer.py --env development --correct-all
  
  # Preview specific correction types
  python weed_visits_analyzer.py --env development --correct-weed-from-visits --preview
  python weed_visits_analyzer.py --env development --correct-visits --preview
  python weed_visits_analyzer.py --env development --correct-visits-from-weed --preview
  
  # Apply WeedLocations corrections to specific fields only
  python weed_visits_analyzer.py --env development --correct-weed-from-visits --fields "Urgency,Status"
"""

import os
import json
import argparse
from datetime import datetime
from tenacity import retry, stop_after_attempt, wait_fixed
from arcgis.gis import GIS
from arcgis.features import FeatureLayer, Table
import pandas as pd


# Global reference date strategy used to determine "latest" visit
# See docstring section "How 'Latest' Visit is Determined" for detailed logic
# This strategy is applied ONCE per weed location in get_latest_visit_per_location()
# Result: Each WeedLocation has ONE "latest" visit, used for ALL field comparisons
REFERENCE_DATE_STRATEGY = 'DateCheck → CreationDate_1'
REFERENCE_DATE_DESCRIPTION = (
  'Latest visit is determined by DateCheck (preferred), '
  'falling back to CreationDate_1 if DateCheck is null'
)

# Initial bulk load date for historical records migrated into the system
# Records created during the initial data migration have this CreationDate_1
# Actual visits occurred prior to this date, so this timestamp doesn't represent
# the true visit creation date and should be ignored in comparisons
# ArcGIS timestamps are stored as epoch milliseconds
# "2021-10-09 21:19:26" UTC = 1633767566000 (epoch ms)
INITIAL_BULK_LOAD_DATE_TIMESTAMP = 1633767566000
INITIAL_BULK_LOAD_DATE_ISO = "2021-10-09 21:19:26"


VISIT_CORRECTION_RULES = [
  {
    'target_field': 'DateCheck',
    'source_field': 'visit_CreationDate_1',
    'display_name': 'DateCheck from CreationDate',
    'condition': lambda row: (
      # Only set DateCheck if it's empty AND CreationDate_1 is not the bulk load date
      pd.isna(row.get('DateCheck')) and
      pd.notna(row.get('visit_CreationDate_1')) and
      row.get('visit_CreationDate_1') != INITIAL_BULK_LOAD_DATE_TIMESTAMP
    ),
    'description': f'Set DateCheck from CreationDate_1 when DateCheck is empty (excludes bulk-loaded records with {INITIAL_BULK_LOAD_DATE_ISO})'
  }
]


VISIT_FROM_WEED_RULES = [
  {
    'visit_field': 'WeedVisitStatus',
    'weed_field': 'ParentStatusWithDomain',
    'display_name': 'Status from WeedLocation',
    'condition': lambda row: (
      # Only copy if WeedVisitStatus is empty AND ParentStatusWithDomain is valid
      pd.isna(row.get('WeedVisitStatus')) and
      pd.notna(row.get('ParentStatusWithDomain')) and
      not str(row.get('ParentStatusWithDomain')).startswith('Purple')
    ),
    'description': 'Copy ParentStatusWithDomain from WeedLocations to WeedVisitStatus when WeedVisitStatus is empty (excludes Purple statuses)'
  }
]


FIELD_COMPARISON_RULES = [
  {
    'weed_field': 'Urgency',
    'visit_field': 'DifficultyChild',
    'display_name': 'Urgency',
    'mismatch_column': 'Urgency_Mismatch',
    'ignore_condition': None,
    'description': 'Urgency level should match DifficultyChild from latest visit'
  },
  {
    'weed_field': 'ParentStatusWithDomain',
    'visit_field': 'WeedVisitStatus',
    'display_name': 'Status',
    'mismatch_column': 'Status_Mismatch',
    'ignore_condition': lambda weed_val, visit_val: pd.notna(weed_val) and str(weed_val).startswith('Purple'),
    'description': 'Parent status should match visit status (ignores Purple statuses)'
  },
  {
    'weed_field': 'DateVisitMadeFromLastVisit',
    'visit_field': 'DateCheck',
    'display_name': 'Date Visit Made',
    'mismatch_column': 'DateVisitMade_Mismatch',
    'ignore_condition': None,
    'description': 'Date of visit should match DateCheck from visit record'
  },
  {
    'weed_field': 'DateForNextVisitFromLastVisit',
    'visit_field': 'DateForReturnVisit',
    'display_name': 'Date For Next Visit',
    'mismatch_column': 'DateForNextVisit_Mismatch',
    'ignore_condition': None,
    'description': 'Next visit date should match DateForReturnVisit from visit record'
  },
  {
    'weed_field': 'LatestVisitStage',
    'visit_field': 'VisitStage',
    'display_name': 'Visit Stage',
    'mismatch_column': 'VisitStage_Mismatch',
    'ignore_condition': None,
    'description': 'Latest visit stage should match VisitStage from visit record'
  },
  {
    'weed_field': 'LatestArea',
    'visit_field': 'Area',
    'display_name': 'Area',
    'mismatch_column': 'Area_Mismatch',
    'ignore_condition': None,
    'description': 'Latest area should match Area from visit record'
  }
  # Commented out - audit fields not currently being synchronized
  # {
  #   'weed_field': 'DateOfLastCreateFromLastVisit',
  #   'visit_field': 'visit_CreationDate_1',
  #   'display_name': 'Creation Date',
  #   'mismatch_column': 'DateOfLastCreate_Mismatch',
  #   'ignore_condition': lambda weed_val, visit_val: (
  #     # Ignore if visit CreationDate_1 is the bulk load date (not a real visit creation date)
  #     visit_val == INITIAL_BULK_LOAD_DATE_TIMESTAMP if pd.notna(visit_val) else False
  #   ),
  #   'description': f'Last creation date should match CreationDate_1 from visit record (ignores bulk-loaded records with creation date {INITIAL_BULK_LOAD_DATE_ISO})',
  #   'category': 'audit'
  # },
  # {
  #   'weed_field': 'DateOfLastEditFromLastVisit',
  #   'visit_field': 'visit_EditDate_1',
  #   'display_name': 'Edit Date',
  #   'mismatch_column': 'DateOfLastEdit_Mismatch',
  #   'ignore_condition': None,
  #   'description': 'Last edit date should match EditDate_1 from visit record',
  #   'category': 'audit'
  # }
]


def query_with_pagination(layer, out_fields, feature_mapper, entity_name="records"):
  """
  Generic pagination function for ArcGIS layers/tables with auto-recovery
  
  Args:
    layer: ArcGIS FeatureLayer or Table object
    out_fields: List of field names to query
    feature_mapper: Function that maps feature.attributes to a dict
    entity_name: Name for logging (e.g., "features", "records")
  
  Returns:
    List of mapped feature dictionaries
  """
  print(f"Loading {entity_name}...")
  
  features_data = []
  offset = 0
  batch_size = 2000
  max_failed_attempts = 5
  failed_batches = []
  
  while True:
    print(f"  Fetching batch at offset {offset}...")
    
    @retry(stop=stop_after_attempt(3), wait=wait_fixed(2))
    def fetch_batch():
      return layer.query(
        where="1=1",
        out_fields=out_fields,
        return_geometry=False,
        result_offset=offset,
        result_record_count=batch_size,
        return_all_records=False
      )
    
    try:
      result = fetch_batch()
      batch_features = result.features
      
      if not batch_features:
        break
      
      features_data.extend([feature_mapper(f.attributes) for f in batch_features])
      print(f"  Loaded {len(batch_features)} {entity_name} (total: {len(features_data)})")
      
      if len(batch_features) < batch_size:
        break
      
      offset += batch_size
      
    except Exception as e:
      print(f"  WARNING: Batch at offset {offset} (size {batch_size}) failed: {type(e).__name__}")
      
      # Try progressively smaller batch sizes
      smaller_sizes = [1000, 500, 250, 100]
      recovered = False
      
      for smaller_size in smaller_sizes:
        if smaller_size >= batch_size:
          continue
        
        print(f"  Retrying with smaller batch size: {smaller_size}")
        sub_offset = offset
        sub_features = []
        
        try:
          while sub_offset < offset + batch_size:
            @retry(stop=stop_after_attempt(2), wait=wait_fixed(1))
            def fetch_small_batch():
              return layer.query(
                where="1=1",
                out_fields=out_fields,
                return_geometry=False,
                result_offset=sub_offset,
                result_record_count=smaller_size,
                return_all_records=False
              )
            
            result = fetch_small_batch()
            batch_features = result.features
            
            if not batch_features:
              break
            
            sub_features.extend([feature_mapper(f.attributes) for f in batch_features])
            sub_offset += smaller_size
            
            if len(batch_features) < smaller_size:
              break
          
          features_data.extend(sub_features)
          print(f"  Recovered {len(sub_features)} {entity_name} with smaller batches (total: {len(features_data)})")
          recovered = True
          break
          
        except Exception as sub_e:
          print(f"  Failed with batch size {smaller_size}: {type(sub_e).__name__}")
          continue
      
      if not recovered:
        print(f"  Could not recover batch at offset {offset}, skipping...")
        failed_batches.append(offset)
        
        if len(failed_batches) >= max_failed_attempts:
          print(f"  ERROR: Too many failed batches ({len(failed_batches)}), stopping.")
          break
      
      offset += batch_size
  
  if failed_batches:
    print(f"\nWARNING: Failed to load {len(failed_batches)} batches at offsets: {failed_batches}")
  
  return features_data


@retry(stop=stop_after_attempt(3), wait=wait_fixed(5))
def connect_arcgis():
  """Connect to ArcGIS Online"""
  username = os.getenv('ARCGIS_USERNAME')
  password = os.getenv('ARCGIS_PASSWORD')
  portal_url = os.getenv('ARCGIS_PORTAL_URL', 'https://www.arcgis.com')
  return GIS(portal_url, username, password)


def get_layers(gis, environment):
  """Get the WeedLocations layer and Visits_Table for the specified environment"""
  script_dir = os.path.dirname(os.path.abspath(__file__))
  env_config_path = os.path.join(script_dir, 'environment_config.json')
  
  with open(env_config_path, 'r') as f:
    env_config = json.load(f)
  
  if environment not in env_config:
    available_envs = list(env_config.keys())
    raise ValueError(f"Environment '{environment}' not found. Available: {available_envs}")
  
  env_settings = env_config[environment]
  weed_layer_id = env_settings['weed_locations_layer_id']
  
  # Get weed layer
  weed_item = gis.content.get(weed_layer_id)
  if not weed_item:
    raise ValueError(f"Could not find WeedLocations layer with ID: {weed_layer_id}")
  weed_layer = FeatureLayer.fromitem(weed_item)
  
  # Get visits table from the feature service
  # The Visits_Table is part of the same feature service as WeedLocations
  if not hasattr(weed_item, 'tables') or len(weed_item.tables) == 0:
    raise ValueError(f"No tables found in WeedLocations feature service: {weed_layer_id}")
  
  # Try to find the Visits_Table by name first
  visits_table = None
  for table in weed_item.tables:
    if 'visit' in table.properties.name.lower():
      visits_table = table
      print(f"Found Visits_Table: {table.properties.name}")
      break
  
  # If not found by name, use the first table (index 0)
  if not visits_table:
    visits_table = weed_item.tables[0]
    print(f"Using first table: {visits_table.properties.name}")
  
  return weed_layer, visits_table


def load_weed_locations(weed_layer):
  """Load all weed locations with relevant fields"""
  out_fields = [
    "OBJECTID", "GlobalID", 
    "Urgency", "ParentStatusWithDomain",
    "DateVisitMadeFromLastVisit", "DateForNextVisitFromLastVisit",
    "LatestVisitStage", "LatestArea",
    "DateOfLastCreateFromLastVisit", "DateOfLastEditFromLastVisit",
    "DateDiscovered", "CreationDate_1"
  ]
  
  def mapper(attrs):
    return {
      'WeedLocation_OBJECTID': attrs.get('OBJECTID'),
      'GlobalID': attrs.get('GlobalID'),
      'Urgency': attrs.get('Urgency'),
      'ParentStatusWithDomain': attrs.get('ParentStatusWithDomain'),
      'DateVisitMadeFromLastVisit': attrs.get('DateVisitMadeFromLastVisit'),
      'DateForNextVisitFromLastVisit': attrs.get('DateForNextVisitFromLastVisit'),
      'LatestVisitStage': attrs.get('LatestVisitStage'),
      'LatestArea': attrs.get('LatestArea'),
      'DateOfLastCreateFromLastVisit': attrs.get('DateOfLastCreateFromLastVisit'),
      'DateOfLastEditFromLastVisit': attrs.get('DateOfLastEditFromLastVisit'),
      'DateDiscovered': attrs.get('DateDiscovered'),
      'weed_CreationDate_1': attrs.get('CreationDate_1')
    }
  
  features_data = query_with_pagination(weed_layer, out_fields, mapper, "WeedLocations features")
  df = pd.DataFrame(features_data)
  
  # Ensure OBJECTID is numeric for consistency
  if 'WeedLocation_OBJECTID' in df.columns:
    df['WeedLocation_OBJECTID'] = pd.to_numeric(df['WeedLocation_OBJECTID'], errors='coerce')
  
  print(f"Loaded {len(df)} weed locations total")
  return df


def load_visits_table(visits_table):
  """Load all visits from Visits_Table"""
  out_fields = [
    "OBJECTID", "GUID_visits",
    "DifficultyChild", "WeedVisitStatus",
    "DateCheck", "DateForReturnVisit",
    "VisitStage", "Area",
    "CreationDate_1", "EditDate_1",
    "VisitDataSource"
  ]
  
  def mapper(attrs):
    return {
      'Visit_OBJECTID': attrs.get('OBJECTID'),
      'GUID_visits': attrs.get('GUID_visits'),
      'DifficultyChild': attrs.get('DifficultyChild'),
      'WeedVisitStatus': attrs.get('WeedVisitStatus'),
      'DateCheck': attrs.get('DateCheck'),
      'DateForReturnVisit': attrs.get('DateForReturnVisit'),
      'VisitStage': attrs.get('VisitStage'),
      'Area': attrs.get('Area'),
      'visit_CreationDate_1': attrs.get('CreationDate_1'),
      'visit_EditDate_1': attrs.get('EditDate_1'),
      'VisitDataSource': attrs.get('VisitDataSource')
    }
  
  visits_data = query_with_pagination(visits_table, out_fields, mapper, "Visits_Table records")
  df = pd.DataFrame(visits_data)
  
  # Ensure OBJECTID is numeric for proper sorting (tiebreaker logic)
  if 'Visit_OBJECTID' in df.columns:
    df['Visit_OBJECTID'] = pd.to_numeric(df['Visit_OBJECTID'], errors='coerce')
  
  print(f"Loaded {len(df)} visit records total")
  return df


def get_latest_visit_per_location(visits_df):
  """
  Get the latest visit for each weed location using global REFERENCE_DATE_STRATEGY
  
  Strategy: DateCheck → CreationDate_1
  - Prefers visits with DateCheck set (sorted by DateCheck descending)
  - Falls back to CreationDate_1 for visits without DateCheck
  
  This "latest" visit is used for ALL field comparisons in FIELD_COMPARISON_RULES
  
  Args:
    visits_df: DataFrame of all visits
  """
  empty_df_columns = [
    'GUID_visits', 'Visit_OBJECTID', 'DifficultyChild', 'WeedVisitStatus',
    'DateCheck', 'DateForReturnVisit', 'VisitStage', 'Area',
    'visit_CreationDate_1', 'visit_EditDate_1', 'VisitDataSource'
  ]
  
  if len(visits_df) == 0:
    return pd.DataFrame(columns=empty_df_columns)
  
  # Get visits with DateCheck (preferred)
  # Use Visit_OBJECTID as tiebreaker when DateCheck values are identical
  visits_with_datecheck = visits_df[visits_df['DateCheck'].notna()].copy()
  
  latest_by_datecheck = (
    visits_with_datecheck.sort_values(['DateCheck', 'Visit_OBJECTID'], ascending=[False, False])
    .groupby('GUID_visits', as_index=False).head(1)
    if len(visits_with_datecheck) > 0
    else pd.DataFrame(columns=empty_df_columns)
  )
  
  # Get visits without DateCheck but with CreationDate_1 (fallback)
  visits_no_datecheck = visits_df[
    visits_df['DateCheck'].isna() & 
    visits_df['visit_CreationDate_1'].notna()
  ].copy()
  
  if len(visits_no_datecheck) > 0:
    # Use Visit_OBJECTID as tiebreaker when CreationDate_1 values are identical
    latest_by_creation = (
      visits_no_datecheck.sort_values(['visit_CreationDate_1', 'Visit_OBJECTID'], ascending=[False, False])
      .groupby('GUID_visits', as_index=False).head(1)
    )
    # Only include GUIDs not already covered by DateCheck
    guids_with_datecheck = set(latest_by_datecheck['GUID_visits'])
    latest_by_creation = latest_by_creation[~latest_by_creation['GUID_visits'].isin(guids_with_datecheck)]
    latest_visits = pd.concat([latest_by_datecheck, latest_by_creation], ignore_index=True)
  else:
    latest_visits = latest_by_datecheck
  
  return latest_visits


def load_weed_locations_with_visits(weed_layer, visits_table):
  """Load weed locations and join with latest visit data"""
  # Load data
  weeds_df = load_weed_locations(weed_layer)
  visits_df = load_visits_table(visits_table)
  
  # Get latest visit per location
  print("Finding latest visit for each weed location...")
  latest_visits = get_latest_visit_per_location(visits_df)
  
  # Build merge columns from rules (all visit fields needed for comparison)
  # Always include audit fields even if not in comparison rules (needed for reference date logic)
  merge_cols = ['GUID_visits', 'Visit_OBJECTID', 'VisitDataSource', 
                'visit_CreationDate_1', 'visit_EditDate_1']
  merge_cols.extend([rule['visit_field'] for rule in FIELD_COMPARISON_RULES])
  # Remove duplicates while preserving order
  merge_cols = list(dict.fromkeys(merge_cols))
  merge_cols = filter_existing_columns(latest_visits, merge_cols)
  
  merged_df = weeds_df.merge(
    latest_visits[merge_cols],
    left_on='GlobalID',
    right_on='GUID_visits',
    how='left'
  )
  
  merged_df = merged_df.drop(columns=['GUID_visits'], errors='ignore')
  
  print(f"Joined data: {len(merged_df)} weed locations with visit information")
  return merged_df


def convert_arcgis_timestamp_to_iso(timestamp):
  """
  Convert ArcGIS epoch millisecond timestamp to ISO 8601 datetime string
  
  Args:
    timestamp: Epoch milliseconds (int or float) or None
  
  Returns:
    ISO formatted datetime string or None if input is None
  """
  if pd.isna(timestamp):
    return None
  
  try:
    # ArcGIS timestamps are in milliseconds since epoch
    dt = datetime.fromtimestamp(timestamp / 1000.0)
    return dt.strftime('%Y-%m-%d %H:%M:%S')
  except (ValueError, TypeError, OSError):
    return None


def convert_date_columns(df, date_columns):
  """Convert multiple date columns in a DataFrame from epoch ms to ISO format"""
  for col in date_columns:
    if col in df.columns:
      df[col] = df[col].apply(convert_arcgis_timestamp_to_iso)
  return df


def convert_correction_dates(corrections_df):
  """
  Convert date values in Old_Value and New_Value columns for corrections log
  Only converts values that are numeric (epoch timestamps), leaves text/other types as-is
  """
  df = corrections_df.copy()
  
  for col in ['Old_Value', 'New_Value']:
    if col in df.columns:
      # Only convert if the value looks like an epoch timestamp (large number)
      df[col] = df[col].apply(lambda x: (
        convert_arcgis_timestamp_to_iso(x) 
        if pd.notna(x) and isinstance(x, (int, float)) and x > 1000000000000 
        else x
      ))
  
  # Always convert Visit_Reference_Date if present
  if 'Visit_Reference_Date' in df.columns:
    df['Visit_Reference_Date'] = df['Visit_Reference_Date'].apply(convert_arcgis_timestamp_to_iso)
  
  return df


def filter_existing_columns(df, column_list):
  """Return only columns that actually exist in the DataFrame"""
  return [col for col in column_list if col in df.columns]


def apply_bold_to_prefixed_cells(sheet, column_names, bold_font, prefix='← '):
  """Apply bold formatting to cells with a specific prefix in given columns"""
  header_row = [cell.value for cell in sheet[1]]
  
  for col_name in column_names:
    if col_name in header_row:
      col_idx = header_row.index(col_name) + 1  # openpyxl uses 1-based indexing
      
      for row_idx in range(2, sheet.max_row + 1):  # Skip header
        cell = sheet.cell(row=row_idx, column=col_idx)
        if cell.value and str(cell.value).startswith(prefix):
          cell.font = bold_font


def get_active_rules(ignore_creation_edit_dates=False, fields_to_include=None):
  """
  Get filtered list of comparison rules based on options
  
  Args:
    ignore_creation_edit_dates: If True, exclude audit category rules
    fields_to_include: Optional list of display names to filter by
  
  Returns:
    List of active comparison rules
  """
  rules = FIELD_COMPARISON_RULES
  
  if ignore_creation_edit_dates:
    rules = [rule for rule in rules if rule.get('category') != 'audit']
  
  if fields_to_include:
    rules = [rule for rule in rules if rule['display_name'] in fields_to_include]
  
  return rules


def apply_batched_updates(layer, updates_by_objectid, corrections_df, objectid_column, 
                          entity_name="record", batch_size=500):
  """
  Apply updates in batches for better performance
  
  Args:
    layer: ArcGIS FeatureLayer or Table to update
    updates_by_objectid: Dict mapping OBJECTID to field updates dict
    corrections_df: DataFrame to track update status
    objectid_column: Column name for OBJECTID in corrections_df
    entity_name: Name for logging (e.g., "record", "location")
    batch_size: Number of features to update per API call
  
  Returns:
    success_count, error_count
  """
  corrections_df['Update_Status'] = 'Pending'
  corrections_df['Update_Error'] = ''
  
  def mark_status(objectids, status, error_msg=''):
    """Mark correction status for list of OBJECTIDs"""
    mask = corrections_df[objectid_column].isin(objectids)
    corrections_df.loc[mask, 'Update_Status'] = status
    if error_msg:
      corrections_df.loc[mask, 'Update_Error'] = error_msg
  
  total_updates = len(updates_by_objectid)
  success_count = 0
  error_count = 0
  
  # Convert to list for batching
  update_items = list(updates_by_objectid.items())
  
  # Process in batches
  for batch_start in range(0, total_updates, batch_size):
    batch_end = min(batch_start + batch_size, total_updates)
    batch_items = update_items[batch_start:batch_end]
    
    print(f"  Updating {entity_name}s {batch_start + 1}-{batch_end} of {total_updates}...")
    
    # Build features for this batch
    features = [
      {'attributes': {'OBJECTID': objectid, **field_updates}}
      for objectid, field_updates in batch_items
    ]
    
    try:
      result = layer.edit_features(updates=features)
      
      if result.get('updateResults'):
        # Process each result in the batch
        for i, update_result in enumerate(result['updateResults']):
          objectid = batch_items[i][0]
          
          if update_result.get('success'):
            success_count += 1
            mark_status([objectid], 'Success')
          else:
            error_count += 1
            error_msg = update_result.get('error', {}).get('description', 'Unknown error')
            print(f"    ERROR {entity_name} OBJECTID {objectid}: {error_msg}")
            mark_status([objectid], 'Failed', error_msg)
      else:
        # Entire batch failed
        error_count += len(batch_items)
        objectids = [item[0] for item in batch_items]
        error_msg = 'No result returned'
        print(f"    ERROR: Batch failed - {error_msg}")
        mark_status(objectids, 'Failed', error_msg)
    
    except Exception as e:
      # Entire batch failed
      error_count += len(batch_items)
      objectids = [item[0] for item in batch_items]
      error_msg = str(e)
      print(f"    ERROR: Batch failed - {error_msg}")
      mark_status(objectids, 'Failed', error_msg)
  
  return success_count, error_count


def check_field_mismatches(merged_df, ignore_creation_edit_dates=False):
  """
  Check all field pairs between WeedLocations and Visits_Table using metadata-driven rules
  
  Args:
    merged_df: DataFrame with merged WeedLocations and Visits data
    ignore_creation_edit_dates: If True, skip checking audit fields (CreationDate_1 and EditDate_1)
  
  Returns DataFrame with mismatch indicators for each field pair and list of active rules
  """
  # Filter rules based on ignore_creation_edit_dates flag
  active_rules = FIELD_COMPARISON_RULES
  if ignore_creation_edit_dates:
    active_rules = [rule for rule in FIELD_COMPARISON_RULES if rule.get('category') != 'audit']
  
  result_df = merged_df.copy()
  mismatch_columns = []
  mismatch_details = []
  
  # Add reference date columns to track which date determined "latest" visit
  # Using global REFERENCE_DATE_STRATEGY: DateCheck → CreationDate_1
  has_visit = result_df['Visit_OBJECTID'].notna()
  result_df['Visit_Reference_Date'] = result_df['DateCheck'].where(
    result_df['DateCheck'].notna(),
    result_df['visit_CreationDate_1']
  )
  result_df['Visit_Reference_Date_Field'] = 'None'
  result_df.loc[has_visit & result_df['DateCheck'].notna(), 'Visit_Reference_Date_Field'] = 'DateCheck'
  result_df.loc[has_visit & result_df['DateCheck'].isna() & result_df['visit_CreationDate_1'].notna(), 'Visit_Reference_Date_Field'] = 'CreationDate_1'
  
  for rule in active_rules:
    weed_field = rule['weed_field']
    visit_field = rule['visit_field']
    mismatch_col = rule['mismatch_column']
    ignore_condition = rule.get('ignore_condition')
    
    # Create mismatch indicator column and reason column
    result_df[mismatch_col] = ''
    result_df[f'{mismatch_col}_Reason'] = ''
    
    # Check for mismatches
    for idx, row in result_df.iterrows():
      weed_val = row.get(weed_field)
      visit_val = row.get(visit_field)
      
      # Skip if no visit record exists at all
      if pd.isna(row.get('Visit_OBJECTID')):
        continue
      
      # Apply custom ignore condition if specified
      if ignore_condition and ignore_condition(weed_val, visit_val):
        continue
      
      # Check for mismatch (includes cases where weed has value but visit is null)
      # Use pandas-safe comparison that handles NaN properly
      is_mismatch = False
      if pd.isna(weed_val) and pd.isna(visit_val):
        is_mismatch = False  # Both null = match
      elif pd.isna(weed_val) or pd.isna(visit_val):
        is_mismatch = True  # One null, one not = mismatch
      else:
        is_mismatch = (weed_val != visit_val)  # Both have values = compare
      
      if is_mismatch:
        result_df.at[idx, mismatch_col] = 'X'
        # Store mismatch reason with actual values
        weed_str = str(weed_val) if pd.notna(weed_val) else 'null'
        visit_str = str(visit_val) if pd.notna(visit_val) else 'null'
        result_df.at[idx, f'{mismatch_col}_Reason'] = (
          f"{rule['display_name']}: WeedLocation={weed_str}, Visit={visit_str}"
        )
    
    mismatch_columns.append(mismatch_col)
  
  # Add a column indicating if ANY field has a mismatch
  result_df['Has_Any_Mismatch'] = (result_df[mismatch_columns] == 'X').any(axis=1)
  
  # Create combined mismatch reason for rows with any mismatch
  reason_cols = [f'{rule["mismatch_column"]}_Reason' for rule in active_rules]
  result_df['Mismatch_Summary'] = result_df[reason_cols].apply(
    lambda row: '; '.join([r for r in row if r]), axis=1
  )
  
  return result_df, mismatch_columns, active_rules


def generate_mismatch_report(merged_df, output_file='weed_visits_field_comparison.xlsx', ignore_creation_edit_dates=False):
  """
  Generate Excel spreadsheet with summary and detailed mismatch data
  
  Args:
    merged_df: DataFrame with merged data
    output_file: Path to output Excel file
    ignore_creation_edit_dates: If True, skip audit fields (CreationDate_1 and EditDate_1)
  """
  # Check field mismatches
  result_df, mismatch_columns, active_rules = check_field_mismatches(merged_df, ignore_creation_edit_dates)
  
  # Calculate summary statistics
  total_locations = len(result_df)
  locations_with_visits = result_df['Visit_OBJECTID'].notna().sum()
  locations_with_mismatches = result_df['Has_Any_Mismatch'].sum()
  
  # Count mismatches per field pair using metadata rules
  field_pair_summary = []
  for rule in active_rules:
    mismatch_col = rule['mismatch_column']
    mismatch_count = (result_df[mismatch_col] == 'X').sum()
    
    field_pair_summary.append({
      'Field': rule['display_name'],
      'WeedLocations_Field': rule['weed_field'],
      'Visits_Table_Field': rule['visit_field'],
      'Latest_Visit_Strategy': REFERENCE_DATE_STRATEGY,
      'Description': rule['description'],
      'Mismatch_Count': mismatch_count,
      'Mismatch_Percentage': f"{(mismatch_count / locations_with_visits * 100):.1f}%" if locations_with_visits > 0 else "0%"
    })
  
  summary_df = pd.DataFrame(field_pair_summary)
  
  # VisitDataSource summary for locations with mismatches
  datasource_summary = []
  if 'VisitDataSource' in result_df.columns:
    mismatched_with_visits = result_df[result_df['Has_Any_Mismatch'] & result_df['Visit_OBJECTID'].notna()]
    
    datasource_counts = mismatched_with_visits['VisitDataSource'].value_counts()
    for datasource, count in datasource_counts.items():
      datasource_summary.append({
        'VisitDataSource': datasource if pd.notna(datasource) else '(Unknown)',
        'Mismatch_Count': count,
        'Percentage_of_Mismatches': f"{(count / locations_with_mismatches * 100):.1f}%" if locations_with_mismatches > 0 else "0%"
      })
    
    # Also add breakdown by field for each datasource
    for rule in active_rules:
      mismatch_col = rule['mismatch_column']
      for datasource in datasource_counts.index:
        ds_field_mismatches = mismatched_with_visits[
          (mismatched_with_visits['VisitDataSource'] == datasource) & 
          (mismatched_with_visits[mismatch_col] == 'X')
        ]
        if len(ds_field_mismatches) > 0:
          datasource_summary.append({
            'VisitDataSource': f"  └─ {datasource if pd.notna(datasource) else '(Unknown)'} - {rule['display_name']}",
            'Mismatch_Count': len(ds_field_mismatches),
            'Percentage_of_Mismatches': f"{(len(ds_field_mismatches) / locations_with_mismatches * 100):.1f}%" if locations_with_mismatches > 0 else "0%"
          })
  
  datasource_summary_df = pd.DataFrame(datasource_summary) if datasource_summary else pd.DataFrame()
  
  # Create overall summary
  overall_summary = pd.DataFrame([
    {'Metric': 'Total WeedLocations', 'Value': total_locations},
    {'Metric': 'Locations with Visit Records', 'Value': locations_with_visits},
    {'Metric': 'Locations with Field Mismatches', 'Value': locations_with_mismatches},
    {'Metric': 'Percentage with Mismatches', 
     'Value': f"{(locations_with_mismatches / locations_with_visits * 100):.1f}%" if locations_with_visits > 0 else "0%"}
  ])
  
  # Filter to only rows with mismatches for detailed sheet
  mismatches_df = result_df[result_df['Has_Any_Mismatch']].copy()
  
  # Build detail columns from active rules
  detail_columns = [
    'WeedLocation_OBJECTID',
    'Visit_OBJECTID',
    'VisitDataSource',
    'Visit_Reference_Date_Field',
    'Visit_Reference_Date',
    'Mismatch_Summary'
  ] + mismatch_columns
  
  # Add field pairs from active rules
  for rule in active_rules:
    detail_columns.extend([rule['weed_field'], rule['visit_field']])
  
  # Ensure all columns exist and convert date columns
  detail_columns = filter_existing_columns(mismatches_df, detail_columns)
  mismatches_detail_df = mismatches_df[detail_columns].copy()
  all_records_df = result_df[detail_columns].copy()
  
  date_columns = [
    'DateVisitMadeFromLastVisit', 'DateCheck',
    'DateForNextVisitFromLastVisit', 'DateForReturnVisit',
    'DateOfLastCreateFromLastVisit', 'visit_CreationDate_1',
    'DateOfLastEditFromLastVisit', 'visit_EditDate_1',
    'DateDiscovered', 'weed_CreationDate_1',
    'Visit_Reference_Date'
  ]
  
  convert_date_columns(mismatches_detail_df, date_columns)
  convert_date_columns(all_records_df, date_columns)
  
  # Prepare data for formatting using active rules
  # Prefix Visit field values with ← where there's a mismatch
  for rule in active_rules:
    weed_field = rule['weed_field']
    visit_field = rule['visit_field']
    mismatch_col = rule['mismatch_column']
    if visit_field in mismatches_detail_df.columns and mismatch_col in mismatches_detail_df.columns:
      # Add prefix to mismatched Visit field values (skip None values)
      mask = (mismatches_detail_df[mismatch_col] == 'X') & (mismatches_detail_df[visit_field].notna())
      mismatches_detail_df.loc[mask, visit_field] = '← ' + mismatches_detail_df.loc[mask, visit_field].astype(str)
    
    # Also prefix in all_records_df for the All Records sheet
    if visit_field in all_records_df.columns and mismatch_col in all_records_df.columns:
      mask = (all_records_df[mismatch_col] == 'X') & (all_records_df[visit_field].notna())
      all_records_df.loc[mask, visit_field] = '← ' + all_records_df.loc[mask, visit_field].astype(str)
  
  # Create Missing Visit Date sheet - visits where DateCheck is not set
  missing_date_df = result_df[
    (result_df['Visit_OBJECTID'].notna()) & 
    (result_df['DateCheck'].isna())
  ].copy()
  
  missing_date_columns = [
    'WeedLocation_OBJECTID', 'Visit_OBJECTID',
    'DateCheck', 'visit_CreationDate_1',
    'WeedVisitStatus', 'DifficultyChild',
    'VisitStage', 'Area'
  ]
  missing_date_columns = filter_existing_columns(missing_date_df, missing_date_columns)
  missing_date_export = missing_date_df[missing_date_columns].copy()
  convert_date_columns(missing_date_export, ['DateCheck', 'visit_CreationDate_1'])
  
  # Prefix CreationDate_1 if not the bulk load date (those are system-generated, not actual visit dates)
  # Note: By this point, dates have been converted to ISO strings, so compare against ISO format
  if 'visit_CreationDate_1' in missing_date_export.columns:
    mask = (missing_date_export['visit_CreationDate_1'].notna()) & \
           (missing_date_export['visit_CreationDate_1'] != INITIAL_BULK_LOAD_DATE_ISO)
    missing_date_export.loc[mask, 'visit_CreationDate_1'] = \
      '← ' + missing_date_export.loc[mask, 'visit_CreationDate_1'].astype(str)
  
  # Create Missing Status sheet - visits where WeedVisitStatus is not set
  # Exclude cases where ParentStatusWithDomain is missing or starts with Purple (per ignore rule)
  missing_status_df = result_df[
    (result_df['Visit_OBJECTID'].notna()) & 
    (result_df['WeedVisitStatus'].isna()) &
    (result_df['ParentStatusWithDomain'].notna()) &
    (~result_df['ParentStatusWithDomain'].astype(str).str.startswith('Purple'))
  ].copy()
  
  missing_status_columns = [
    'WeedLocation_OBJECTID', 'Visit_OBJECTID',
    'WeedVisitStatus', 'ParentStatusWithDomain',
    'DateCheck', 'visit_CreationDate_1',
    'DifficultyChild', 'VisitStage', 'Area'
  ]
  missing_status_columns = filter_existing_columns(missing_status_df, missing_status_columns)
  missing_status_export = missing_status_df[missing_status_columns].copy()
  convert_date_columns(missing_status_export, ['DateCheck', 'visit_CreationDate_1'])
  
  # Write to Excel
  print(f"\nGenerating report: {output_file}")
  with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    overall_summary.to_excel(writer, sheet_name='Overall Summary', index=False)
    summary_df.to_excel(writer, sheet_name='Field Pair Summary', index=False)
    
    # Add VisitDataSource summary if available
    if not datasource_summary_df.empty:
      datasource_summary_df.to_excel(writer, sheet_name='Mismatch by DataSource', index=False)
    
    mismatches_detail_df.to_excel(writer, sheet_name='Detailed Mismatches', index=False)
    missing_date_export.to_excel(writer, sheet_name='Missing Visit Date', index=False)
    missing_status_export.to_excel(writer, sheet_name='Missing Status', index=False)
    
    # Also include full data for reference
    all_records_df.to_excel(writer, sheet_name='All Records', index=False)
    
    # Apply bold formatting to Visit fields with mismatches
    from openpyxl.styles import Font
    
    workbook = writer.book
    bold_font = Font(bold=True)
    
    # Get list of visit fields from active rules
    visit_fields = [rule['visit_field'] for rule in active_rules]
    
    # Format both Detailed Mismatches and All Records sheets
    for sheet_name in ['Detailed Mismatches', 'All Records']:
      if sheet_name in workbook.sheetnames:
        apply_bold_to_prefixed_cells(workbook[sheet_name], visit_fields, bold_font)
    
    # Format Missing Visit Date sheet
    if 'Missing Visit Date' in workbook.sheetnames:
      apply_bold_to_prefixed_cells(workbook['Missing Visit Date'], ['visit_CreationDate_1'], bold_font)
  
  print(f"Report saved to: {output_file}")
  print(f"\nSummary:")
  print(f"  Total locations: {total_locations:,}")
  print(f"  Locations with visits: {locations_with_visits:,}")
  print(f"  Locations with mismatches: {locations_with_mismatches:,}")
  if locations_with_visits > 0:
    print(f"  Mismatch rate: {(locations_with_mismatches / locations_with_visits * 100):.1f}%")
  print(f"  Missing visit date (DateCheck): {len(missing_date_export):,}")
  print(f"  Missing status (WeedVisitStatus): {len(missing_status_export):,}")
  
  # Print VisitDataSource breakdown if available
  if not datasource_summary_df.empty:
    print("\nMismatches by VisitDataSource:")
    for _, row in datasource_summary_df.iterrows():
      print(f"  {row['VisitDataSource']}: {row['Mismatch_Count']} ({row['Percentage_of_Mismatches']})")
  
  return overall_summary, summary_df, mismatches_detail_df


def analyze_date_matching(merged_df):
  """
  Analyze date matching between WeedLocations and Visits_Table
  
  Returns counts for various date matching scenarios
  """
  total = len(merged_df)
  
  # Category 1: DateVisitMadeFromLastVisit matches latest visit DateCheck
  matching_datecheck = merged_df[
    (merged_df['DateVisitMadeFromLastVisit'].notna()) &
    (merged_df['DateCheck'].notna()) &
    (merged_df['DateVisitMadeFromLastVisit'] == merged_df['DateCheck'])
  ]
  
  # Category 2: Weed date set, visit DateCheck not set
  weed_set_visit_datecheck_not = merged_df[
    (merged_df['DateVisitMadeFromLastVisit'].notna()) &
    (merged_df['DateCheck'].isna())
  ]
  
  # Category 2a: Of those, how many match visit Creation_Date1?
  weed_matches_visit_creation = weed_set_visit_datecheck_not[
    (weed_set_visit_datecheck_not['visit_CreationDate_1'].notna()) &
    (weed_set_visit_datecheck_not['DateVisitMadeFromLastVisit'] == weed_set_visit_datecheck_not['visit_CreationDate_1'])
  ]
  
  # Category 3: Weed date not set
  weed_not_set = merged_df[merged_df['DateVisitMadeFromLastVisit'].isna()]
  
  # Category 4: Both DateCheck set but don't match
  not_matching = merged_df[
    (merged_df['DateVisitMadeFromLastVisit'].notna()) &
    (merged_df['DateCheck'].notna()) &
    (merged_df['DateVisitMadeFromLastVisit'] != merged_df['DateCheck'])
  ]
  
  # New analysis: Date source hierarchy
  # Count features by which date field is set (priority order)
  has_visit_datecheck = merged_df[merged_df['DateCheck'].notna()]
  has_visit_creation_only = merged_df[
    (merged_df['DateCheck'].isna()) &
    (merged_df['visit_CreationDate_1'].notna())
  ]
  has_weed_discovered_only = merged_df[
    (merged_df['DateCheck'].isna()) &
    (merged_df['visit_CreationDate_1'].isna()) &
    (merged_df['DateDiscovered'].notna())
  ]
  has_weed_creation_only = merged_df[
    (merged_df['DateCheck'].isna()) &
    (merged_df['visit_CreationDate_1'].isna()) &
    (merged_df['DateDiscovered'].isna()) &
    (merged_df['weed_CreationDate_1'].notna())
  ]
  has_no_dates = merged_df[
    (merged_df['DateCheck'].isna()) &
    (merged_df['visit_CreationDate_1'].isna()) &
    (merged_df['DateDiscovered'].isna()) &
    (merged_df['weed_CreationDate_1'].isna())
  ]
  
  return {
    'total': total,
    'matching_datecheck': len(matching_datecheck),
    'weed_set_visit_datecheck_not': len(weed_set_visit_datecheck_not),
    'weed_matches_visit_creation': len(weed_matches_visit_creation),
    'weed_not_set': len(weed_not_set),
    'not_matching': len(not_matching),
    # Date source hierarchy
    'has_visit_datecheck': len(has_visit_datecheck),
    'has_visit_creation_only': len(has_visit_creation_only),
    'has_weed_discovered_only': len(has_weed_discovered_only),
    'has_weed_creation_only': len(has_weed_creation_only),
    'has_no_dates': len(has_no_dates)
  }


def format_report(counts):
  """Format analysis results as a readable report"""
  total = counts['total']
  
  report = []
  report.append("\n" + "=" * 80)
  report.append("WEED LOCATIONS vs VISITS TABLE - DATE ANALYSIS REPORT")
  report.append("=" * 80)
  report.append(f"\nTotal Features Analyzed: {total:,}")
  report.append("\n" + "-" * 80)
  report.append("PART 1: DATE SYNCHRONIZATION ANALYSIS")
  report.append("-" * 80)
  
  if total == 0:
    report.append("\nNo features found to analyze.")
    report.append("=" * 80)
    return "\n".join(report)
  
  # Category 1: Matching DateCheck
  matching = counts['matching_datecheck']
  matching_pct = (matching / total * 100) if total > 0 else 0
  report.append(f"\n1. DateVisitMadeFromLastVisit == Latest Visit DateCheck")
  report.append(f"   Count: {matching:,} ({matching_pct:.1f}%)")
  report.append(f"   Status: ✓ Dates are synchronized with visit DateCheck")
  
  # Category 2: Weed set, visit DateCheck not set
  weed_set = counts['weed_set_visit_datecheck_not']
  weed_set_pct = (weed_set / total * 100) if total > 0 else 0
  weed_matches_creation = counts['weed_matches_visit_creation']
  weed_matches_creation_pct = (weed_matches_creation / weed_set * 100) if weed_set > 0 else 0
  report.append(f"\n2. DateVisitMadeFromLastVisit SET, but Latest Visit DateCheck NOT SET")
  report.append(f"   Count: {weed_set:,} ({weed_set_pct:.1f}%)")
  report.append(f"   Status: ⚠ Weed location has date but no related visit DateCheck")
  report.append(f"   → Of these, {weed_matches_creation:,} ({weed_matches_creation_pct:.1f}%) match")
  report.append(f"     Latest Visit Creation_Date1")
  
  # Category 3: Weed not set
  weed_not_set = counts['weed_not_set']
  weed_not_set_pct = (weed_not_set / total * 100) if total > 0 else 0
  report.append(f"\n3. DateVisitMadeFromLastVisit NOT SET")
  report.append(f"   Count: {weed_not_set:,} ({weed_not_set_pct:.1f}%)")
  report.append(f"   Status: ℹ Weed location has no visit date recorded")
  
  # Category 4: Not matching
  not_matching = counts['not_matching']
  not_matching_pct = (not_matching / total * 100) if total > 0 else 0
  report.append(f"\n4. DateVisitMadeFromLastVisit != Latest Visit DateCheck")
  report.append(f"   Count: {not_matching:,} ({not_matching_pct:.1f}%)")
  report.append(f"   Status: ✗ Dates are out of sync")
  
  # Part 2: Date Source Hierarchy
  report.append("\n" + "-" * 80)
  report.append("PART 2: DATE SOURCE HIERARCHY (Priority Order)")
  report.append("-" * 80)
  
  has_visit_datecheck = counts['has_visit_datecheck']
  has_visit_datecheck_pct = (has_visit_datecheck / total * 100) if total > 0 else 0
  report.append(f"\n1. Latest Visit DateCheck is SET")
  report.append(f"   Count: {has_visit_datecheck:,} ({has_visit_datecheck_pct:.1f}%)")
  report.append(f"   Priority: Highest - Use this date")
  
  has_visit_creation = counts['has_visit_creation_only']
  has_visit_creation_pct = (has_visit_creation / total * 100) if total > 0 else 0
  report.append(f"\n2. Latest Visit Creation_Date1 is SET (DateCheck not set)")
  report.append(f"   Count: {has_visit_creation:,} ({has_visit_creation_pct:.1f}%)")
  report.append(f"   Priority: High - Fallback to creation date")
  
  has_weed_discovered = counts['has_weed_discovered_only']
  has_weed_discovered_pct = (has_weed_discovered / total * 100) if total > 0 else 0
  report.append(f"\n3. Weed DateDiscovered is SET (no visit dates)")
  report.append(f"   Count: {has_weed_discovered:,} ({has_weed_discovered_pct:.1f}%)")
  report.append(f"   Priority: Medium - Use weed discovery date")
  
  has_weed_creation = counts['has_weed_creation_only']
  has_weed_creation_pct = (has_weed_creation / total * 100) if total > 0 else 0
  report.append(f"\n4. Weed Creation_Date1 is SET (no other dates)")
  report.append(f"   Count: {has_weed_creation:,} ({has_weed_creation_pct:.1f}%)")
  report.append(f"   Priority: Low - Last resort date")
  
  has_no_dates = counts['has_no_dates']
  has_no_dates_pct = (has_no_dates / total * 100) if total > 0 else 0
  report.append(f"\n5. NO DATES SET")
  report.append(f"   Count: {has_no_dates:,} ({has_no_dates_pct:.1f}%)")
  report.append(f"   Priority: None - No date information available")
  
  # Summary
  report.append("\n" + "-" * 80)
  report.append("SUMMARY")
  report.append("-" * 80)
  in_sync = matching
  out_of_sync = weed_set + not_matching
  no_data = weed_not_set
  
  report.append(f"In Sync:     {in_sync:,} ({in_sync/total*100:.1f}%)")
  report.append(f"Out of Sync: {out_of_sync:,} ({out_of_sync/total*100:.1f}%)")
  report.append(f"No Data:     {no_data:,} ({no_data/total*100:.1f}%)")
  report.append("=" * 80 + "\n")
  
  return "\n".join(report)


def correct_visits_table(visits_table, visits_df, preview_only=False):
  """
  Correct Visit_Table fields using VISIT_CORRECTION_RULES
  
  Args:
    visits_table: Visits_Table object
    visits_df: DataFrame with ALL visit records
    preview_only: If True, show what would change without making changes
  
  Returns:
    DataFrame with correction details
  """
  print("\n" + "=" * 80)
  if preview_only:
    print("PREVIEW MODE: Showing Visit corrections that would be made (no actual updates)")
  else:
    print("CORRECTION MODE: Updating Visits_Table based on correction rules")
  print("=" * 80 + "\n")
  
  corrections = []
  
  for rule in VISIT_CORRECTION_RULES:
    target_field = rule['target_field']
    source_field = rule['source_field']
    condition = rule['condition']
    display_name = rule['display_name']
    
    # Find records that meet the correction condition (latest visits only)
    applicable_records = latest_visits_df[latest_visits_df.apply(condition, axis=1)].copy()
    
    if len(applicable_records) == 0:
      continue
    
    print(f"Processing {display_name} ({target_field} ← {source_field}): {len(applicable_records)} records")
    
    for idx, row in applicable_records.iterrows():
      visit_objectid = row['Visit_OBJECTID']
      source_val = row.get(source_field)
      
      corrections.append({
        'Visit_OBJECTID': visit_objectid,
        'Rule': display_name,
        'Target_Field': target_field,
        'Source_Field': source_field,
        'New_Value': source_val,
        'GUID_visits': row.get('GUID_visits')
      })
  
  corrections_df = pd.DataFrame(corrections)
  
  if len(corrections_df) == 0:
    print("\nNo visit corrections to apply")
    return corrections_df
  
  print(f"\nTotal visit corrections to apply: {len(corrections_df)}")
  
  if preview_only:
    print("\n[PREVIEW ONLY - No changes made]")
    return corrections_df
  
  # Group by OBJECTID
  updates_by_objectid = {}
  for _, correction in corrections_df.iterrows():
    objectid = correction['Visit_OBJECTID']
    if objectid not in updates_by_objectid:
      updates_by_objectid[objectid] = {}
    
    field_name = correction['Target_Field']
    new_value = correction['New_Value']
    
    # Convert pandas NaN to Python None for ArcGIS compatibility
    if pd.isna(new_value):
      new_value = None
    
    updates_by_objectid[objectid][field_name] = new_value
  
  # Apply corrections in batches
  print("\nApplying visit corrections in batches...")
  success_count, error_count = apply_batched_updates(
    visits_table, 
    updates_by_objectid, 
    corrections_df,
    'Visit_OBJECTID',
    'visit',
    batch_size=500
  )
  
  print(f"\nVisit corrections complete:")
  print(f"  Successfully updated: {success_count} records")
  print(f"  Errors: {error_count} records")
  
  return corrections_df


def correct_visits_from_weed(visits_table, merged_df, preview_only=False):
  """
  Update latest Visit fields based on WeedLocations data using VISIT_FROM_WEED_RULES
  
  Args:
    visits_table: Visits_Table object
    merged_df: DataFrame with merged WeedLocations and latest Visit data
    preview_only: If True, show what would change without making changes
  
  Returns:
    DataFrame with correction details
  """
  print("\n" + "=" * 80)
  if preview_only:
    print("PREVIEW MODE: Showing Visit updates from WeedLocations (no actual updates)")
  else:
    print("CORRECTION MODE: Updating latest Visits from WeedLocations data")
  print("=" * 80 + "\n")
  
  corrections = []
  
  for rule in VISIT_FROM_WEED_RULES:
    visit_field = rule['visit_field']
    weed_field = rule['weed_field']
    condition = rule['condition']
    display_name = rule['display_name']
    
    # Find records that meet the correction condition
    applicable_records = merged_df[merged_df.apply(condition, axis=1)].copy()
    
    if len(applicable_records) == 0:
      continue
    
    print(f"Processing {display_name} ({visit_field} ← {weed_field}): {len(applicable_records)} records")
    
    for idx, row in applicable_records.iterrows():
      visit_objectid = row.get('Visit_OBJECTID')
      
      # Skip if no visit record exists
      if pd.isna(visit_objectid):
        continue
      
      weed_val = row.get(weed_field)
      
      corrections.append({
        'Visit_OBJECTID': visit_objectid,
        'WeedLocation_OBJECTID': row.get('WeedLocation_OBJECTID'),
        'Rule': display_name,
        'Visit_Field': visit_field,
        'Source_Field': weed_field,
        'New_Value': weed_val,
        'GUID_visits': row.get('GUID_visits')
      })
  
  corrections_df = pd.DataFrame(corrections)
  
  if len(corrections_df) == 0:
    print("\nNo visit-from-weed corrections to apply")
    return corrections_df
  
  print(f"\nTotal visit-from-weed corrections to apply: {len(corrections_df)}")
  
  if preview_only:
    print("\n[PREVIEW ONLY - No changes made]")
    return corrections_df
  
  # Group by Visit OBJECTID
  updates_by_objectid = {}
  for _, correction in corrections_df.iterrows():
    objectid = correction['Visit_OBJECTID']
    if objectid not in updates_by_objectid:
      updates_by_objectid[objectid] = {}
    
    field_name = correction['Visit_Field']
    new_value = correction['New_Value']
    
    # Convert pandas NaN to Python None for ArcGIS compatibility
    if pd.isna(new_value):
      new_value = None
    
    updates_by_objectid[objectid][field_name] = new_value
  
  # Apply corrections in batches
  print("\nApplying visit-from-weed corrections in batches...")
  success_count, error_count = apply_batched_updates(
    visits_table, 
    updates_by_objectid, 
    corrections_df,
    'Visit_OBJECTID',
    'visit',
    batch_size=500
  )
  
  print(f"\nVisit-from-weed corrections complete:")
  print(f"  Successfully updated: {success_count} records")
  print(f"  Errors: {error_count} records")
  
  return corrections_df


def correct_mismatches(weed_layer, merged_df, ignore_creation_edit_dates=False, 
                       fields_to_correct=None, preview_only=False):
  """
  Correct field mismatches by updating WeedLocations to match Visit data
  
  Args:
    weed_layer: WeedLocations FeatureLayer
    merged_df: DataFrame with merged data and mismatch indicators
    ignore_creation_edit_dates: If True, skip audit fields
    fields_to_correct: List of field names to correct (None = all)
    preview_only: If True, show what would change without making changes
  
  Returns:
    DataFrame with correction details
  """
  print("\n" + "=" * 80)
  if preview_only:
    print("PREVIEW MODE: Showing changes that would be made (no actual updates)")
  else:
    print("CORRECTION MODE: Updating WeedLocations to match Visit data")
  print("=" * 80 + "\n")
  
  active_rules = get_active_rules(ignore_creation_edit_dates, fields_to_correct)
  
  # Find records with mismatches
  mismatched_df = merged_df[merged_df.get('Has_Any_Mismatch', False)].copy()
  
  if len(mismatched_df) == 0:
    print("No mismatches found. Nothing to correct.")
    return pd.DataFrame()
  
  print(f"Found {len(mismatched_df)} records with mismatches")
  print(f"Correction rules active: {len(active_rules)}")
  print()
  
  corrections = []
  ignored_summary = []
  
  for rule in active_rules:
    weed_field = rule['weed_field']
    visit_field = rule['visit_field']
    mismatch_col = rule['mismatch_column']
    display_name = rule['display_name']
    ignore_condition = rule.get('ignore_condition')
    
    # Find records with this specific mismatch
    field_mismatches = mismatched_df[mismatched_df[mismatch_col] == 'X'].copy()
    
    if len(field_mismatches) == 0:
      continue
    
    ignored_count = 0
    null_count = 0
    
    print(f"Processing {display_name} ({weed_field} ← {visit_field}): {len(field_mismatches)} mismatches")
    
    for idx, row in field_mismatches.iterrows():
      weed_objectid = row['WeedLocation_OBJECTID']
      weed_val = row.get(weed_field)
      visit_val = row.get(visit_field)
      
      # Skip if ignore condition applies
      if ignore_condition and ignore_condition(weed_val, visit_val):
        ignored_count += 1
        continue
      
      # Count null visit values but still correct them (clear weed field to match)
      if pd.isna(visit_val):
        null_count += 1
      
      corrections.append({
        'WeedLocation_OBJECTID': weed_objectid,
        'Visit_OBJECTID': row.get('Visit_OBJECTID'),
        'Field': display_name,
        'WeedLocations_Field': weed_field,
        'Old_Value': weed_val,
        'New_Value': visit_val,  # This may be null/None - will clear the weed field
        'Visit_Reference_Date_Field': row.get('Visit_Reference_Date_Field'),
        'Visit_Reference_Date': row.get('Visit_Reference_Date'),
        'VisitDataSource': row.get('VisitDataSource')
      })
    
    if ignored_count > 0 or null_count > 0:
      ignored_summary.append({
        'Field': display_name,
        'Total_Mismatches': len(field_mismatches),
        'Ignored_by_Condition': ignored_count,
        'Null_Visit_Values': null_count,
        'Will_Correct': len(field_mismatches) - ignored_count - null_count
      })
      print(f"  → Will correct: {len(field_mismatches) - ignored_count} (ignored: {ignored_count}, clearing to null: {null_count})")
  
  corrections_df = pd.DataFrame(corrections)
  
  if len(corrections_df) == 0:
    print("\nNo corrections to apply (all mismatches have null visit values or are ignored)")
    return corrections_df
  
  print(f"\nTotal corrections to apply: {len(corrections_df)}")
  
  if ignored_summary:
    print("\nMismatch breakdown (why some won't be corrected):")
    ignored_df = pd.DataFrame(ignored_summary)
    for _, row in ignored_df.iterrows():
      print(f"  {row['Field']}: {row['Total_Mismatches']} total → {row['Will_Correct']} will correct")
      if row['Ignored_by_Condition'] > 0:
        print(f"    - {row['Ignored_by_Condition']} ignored by rule condition (e.g., Purple status, bulk load date)")
      if row['Null_Visit_Values'] > 0:
        print(f"    - {row['Null_Visit_Values']} will be cleared to null (latest visit has no value)")
  
  if preview_only:
    print("\n[PREVIEW ONLY - No changes made]")
    return corrections_df
  
  # Group by OBJECTID to minimize update calls
  updates_by_objectid = {}
  for _, correction in corrections_df.iterrows():
    objectid = correction['WeedLocation_OBJECTID']
    if objectid not in updates_by_objectid:
      updates_by_objectid[objectid] = {}
    
    field_name = correction['WeedLocations_Field']
    new_value = correction['New_Value']
    
    # Convert pandas NaN to Python None for ArcGIS compatibility
    if pd.isna(new_value):
      new_value = None
    
    updates_by_objectid[objectid][field_name] = new_value
  
  # Apply corrections in batches
  print("\nApplying corrections in batches...")
  success_count, error_count = apply_batched_updates(
    weed_layer,
    updates_by_objectid,
    corrections_df,
    'WeedLocation_OBJECTID',
    'location',
    batch_size=500
  )
  
  print(f"\nCorrections complete:")
  print(f"  Successfully updated: {success_count} records")
  print(f"  Errors: {error_count} records")
  
  # Show summary by field
  print("\nUpdate summary by field:")
  for field in corrections_df['Field'].unique():
    field_corrections = corrections_df[corrections_df['Field'] == field]
    successful = len(field_corrections[field_corrections['Update_Status'] == 'Success'])
    failed = len(field_corrections[field_corrections['Update_Status'] == 'Failed'])
    print(f"  {field}: {successful} successful, {failed} failed")
  
  return corrections_df


def analyze_weed_visits(environment, output_file=None, ignore_creation_edit_dates=False,
                       correct_mismatches_flag=False, fields_to_correct=None, 
                       correct_visits_flag=False, correct_visits_from_weed_flag=False,
                       preview_only=False):
  """
  Main analysis function
  
  Args:
    environment: Environment name (e.g., 'production', 'development')
    output_file: Optional output Excel file path
    ignore_creation_edit_dates: If True, skip CreationDate_1 and EditDate_1 comparisons
    correct_mismatches_flag: If True, apply corrections to WeedLocations
    fields_to_correct: List of field names to correct (None = all)
    correct_visits_flag: If True, apply VISIT_CORRECTION_RULES to all visits
    correct_visits_from_weed_flag: If True, apply VISIT_FROM_WEED_RULES to latest visits
    preview_only: If True, show what would change without making changes
  """
  print(f"Starting Weed Visits Analysis for '{environment}' environment...")
  
  if ignore_creation_edit_dates:
    print("Note: Ignoring CreationDate_1 and EditDate_1 field comparisons")
  
  # Connect to ArcGIS
  gis = connect_arcgis()
  
  # Get layers
  weed_layer, visits_table = get_layers(gis, environment)
  
  # Load data (before merging, in case we need to correct visits)
  weeds_df = load_weed_locations(weed_layer)
  visits_df = load_visits_table(visits_table)
  
  # Apply visit corrections first if requested
  visit_corrections_df = None
  if correct_visits_flag:
    visit_corrections_df = correct_visits_table(visits_table, visits_df, preview_only)
    
    # Save visit corrections log
    if visit_corrections_df is not None and len(visit_corrections_df) > 0:
      # Convert date values to ISO format for readability
      visit_corrections_export = convert_correction_dates(visit_corrections_df)
      
      timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
      mode = 'preview' if preview_only else 'applied'
      visit_corrections_file = f'visit_corrections_{mode}_{environment}_{timestamp}.xlsx'
      visit_corrections_export.to_excel(visit_corrections_file, index=False)
      print(f"Visit corrections log saved to: {visit_corrections_file}")
    
    # Reload visits if we actually made changes
    # IMPORTANT: Must recompute which visit is "latest" because DateCheck values may have changed!
    if not preview_only and visit_corrections_df is not None and len(visit_corrections_df) > 0:
      successful = visit_corrections_df[visit_corrections_df['Update_Status'] == 'Success']
      if len(successful) > 0:
        print("\nReloading visits and recomputing latest visits after visit corrections...")
        visits_df = load_visits_table(visits_table)
        latest_visits = get_latest_visit_per_location(visits_df)
        
        # Rebuild merged data with new latest visits
        merge_cols = ['GUID_visits', 'Visit_OBJECTID', 'VisitDataSource',
                      'visit_CreationDate_1', 'visit_EditDate_1']
        merge_cols.extend([rule['visit_field'] for rule in FIELD_COMPARISON_RULES])
        merge_cols = list(dict.fromkeys(merge_cols))
        merge_cols = filter_existing_columns(latest_visits, merge_cols)
        
        merged_df = weeds_df.merge(
          latest_visits[merge_cols],
          left_on='GlobalID',
          right_on='GUID_visits',
          how='left'
        )
  
  # Now merge for analysis
  print("\nMerging WeedLocations with latest visit data...")
  latest_visits = get_latest_visit_per_location(visits_df)
  # Build merge columns from rules (all visit fields needed for comparison)
  # Always include audit fields even if not in comparison rules (needed for reference date logic)
  merge_cols = ['GUID_visits', 'Visit_OBJECTID', 'VisitDataSource', 
                'visit_CreationDate_1', 'visit_EditDate_1']
  merge_cols.extend([rule['visit_field'] for rule in FIELD_COMPARISON_RULES])
  # Remove duplicates while preserving order
  merge_cols = list(dict.fromkeys(merge_cols))
  merge_cols = filter_existing_columns(latest_visits, merge_cols)
  
  merged_df = weeds_df.merge(
    latest_visits[merge_cols],
    left_on='GlobalID',
    right_on='GUID_visits',
    how='left'
  )
  
  # Ensure audit fields are present (in case they were filtered out or missing from source)
  if 'visit_CreationDate_1' not in merged_df.columns:
    merged_df['visit_CreationDate_1'] = None
  if 'visit_EditDate_1' not in merged_df.columns:
    merged_df['visit_EditDate_1'] = None
  
  # Apply visit-from-weed corrections if requested (updates latest visits from WeedLocations)
  visit_from_weed_corrections_df = None
  if correct_visits_from_weed_flag:
    visit_from_weed_corrections_df = correct_visits_from_weed(visits_table, merged_df, preview_only)
    
    # Save corrections log
    if visit_from_weed_corrections_df is not None and len(visit_from_weed_corrections_df) > 0:
      visit_from_weed_export = convert_correction_dates(visit_from_weed_corrections_df)
      
      timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
      mode = 'preview' if preview_only else 'applied'
      vfw_corrections_file = f'visit_from_weed_corrections_{mode}_{environment}_{timestamp}.xlsx'
      visit_from_weed_export.to_excel(vfw_corrections_file, index=False)
      print(f"Visit-from-weed corrections log saved to: {vfw_corrections_file}")
    
    # Reload and re-merge if we actually made changes
    # IMPORTANT: Recompute "latest" in case earlier steps changed DateCheck values!
    if not preview_only and visit_from_weed_corrections_df is not None and len(visit_from_weed_corrections_df) > 0:
      successful = visit_from_weed_corrections_df[visit_from_weed_corrections_df['Update_Status'] == 'Success']
      if len(successful) > 0:
        print("\nReloading visits and recomputing latest visits after visit-from-weed corrections...")
        visits_df = load_visits_table(visits_table)
        latest_visits = get_latest_visit_per_location(visits_df)
        merge_cols = ['GUID_visits', 'Visit_OBJECTID', 'VisitDataSource',
                      'visit_CreationDate_1', 'visit_EditDate_1']
        merge_cols.extend([rule['visit_field'] for rule in FIELD_COMPARISON_RULES])
        merge_cols = list(dict.fromkeys(merge_cols))
        merge_cols = filter_existing_columns(latest_visits, merge_cols)
        
        merged_df = weeds_df.merge(
          latest_visits[merge_cols],
          left_on='GlobalID',
          right_on='GUID_visits',
          how='left'
        )
        
        # Ensure audit fields are present
        if 'visit_CreationDate_1' not in merged_df.columns:
          merged_df['visit_CreationDate_1'] = None
        if 'visit_EditDate_1' not in merged_df.columns:
          merged_df['visit_EditDate_1'] = None
  
  # Generate field comparison report
  if output_file is None:
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'weed_visits_field_comparison_{environment}_{timestamp}.xlsx'
  
  print("\nAnalyzing field mismatches...")
  overall_summary, field_summary, mismatches = generate_mismatch_report(
    merged_df, output_file, ignore_creation_edit_dates
  )
  
  # Apply corrections if requested
  corrections_df = None
  if correct_mismatches_flag:
    # Re-check for mismatches using the result_df from check_field_mismatches
    result_df, mismatch_columns, active_rules = check_field_mismatches(merged_df, ignore_creation_edit_dates)
    corrections_df = correct_mismatches(
      weed_layer, 
      result_df, 
      ignore_creation_edit_dates,
      fields_to_correct,
      preview_only
    )
    
    # Save corrections log if any were made
    if corrections_df is not None and len(corrections_df) > 0:
      # Convert date values to ISO format for readability
      corrections_export = convert_correction_dates(corrections_df)
      
      timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
      mode = 'preview' if preview_only else 'applied'
      corrections_file = f'weed_visits_corrections_{mode}_{environment}_{timestamp}.xlsx'
      corrections_export.to_excel(corrections_file, index=False)
      print(f"\nCorrections log saved to: {corrections_file}")
    
    # Reload and regenerate report after applying corrections (if not preview)
    if not preview_only and corrections_df is not None and len(corrections_df) > 0:
      successful_updates = corrections_df[corrections_df['Update_Status'] == 'Success']
      if len(successful_updates) > 0:
        print("\nReloading all data and regenerating report after corrections...")
        
        # Reload both weeds AND visits (visits may have been updated by earlier correction steps)
        weeds_df = load_weed_locations(weed_layer)
        visits_df = load_visits_table(visits_table)
        
        # Recompute latest visits
        latest_visits = get_latest_visit_per_location(visits_df)
        
        # Rebuild merge columns
        merge_cols = ['GUID_visits', 'Visit_OBJECTID', 'VisitDataSource',
                      'visit_CreationDate_1', 'visit_EditDate_1']
        merge_cols.extend([rule['visit_field'] for rule in FIELD_COMPARISON_RULES])
        merge_cols = list(dict.fromkeys(merge_cols))
        merge_cols = filter_existing_columns(latest_visits, merge_cols)
        
        # Re-merge with fresh data
        merged_df = weeds_df.merge(
          latest_visits[merge_cols],
          left_on='GlobalID',
          right_on='GUID_visits',
          how='left'
        )
        
        # Ensure audit fields are present
        if 'visit_CreationDate_1' not in merged_df.columns:
          merged_df['visit_CreationDate_1'] = None
        if 'visit_EditDate_1' not in merged_df.columns:
          merged_df['visit_EditDate_1'] = None
        
        # Regenerate report with updated data
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        post_correction_file = f'weed_visits_field_comparison_{environment}_after_corrections_{timestamp}.xlsx'
        print(f"\nGenerating post-correction analysis report...")
        overall_summary, field_summary, mismatches = generate_mismatch_report(
          merged_df, post_correction_file, ignore_creation_edit_dates
        )
        print(f"Post-correction report saved to: {post_correction_file}")
  
  # Print final summary
  print("\n" + "=" * 80)
  print("FINAL SUMMARY")
  print("=" * 80)
  
  # Get mismatch counts by field
  result_df, mismatch_columns, active_rules = check_field_mismatches(merged_df, ignore_creation_edit_dates)
  
  print(f"\nTotal WeedLocations: {len(result_df):,}")
  print(f"Locations with visits: {result_df['Visit_OBJECTID'].notna().sum():,}")
  print(f"Locations with field mismatches: {result_df['Has_Any_Mismatch'].sum():,}")
  
  # Show mismatch breakdown by field
  print("\nMismatch breakdown by field:")
  for rule in active_rules:
    mismatch_col = rule['mismatch_column']
    mismatch_count = (result_df[mismatch_col] == 'X').sum()
    if mismatch_count > 0:
      print(f"  {rule['display_name']}: {mismatch_count:,}")
  
  # Show correction summary if corrections were applied
  if corrections_df is not None and len(corrections_df) > 0:
    print(f"\nWeedLocations corrections applied: {len(corrections_df[corrections_df['Update_Status'] == 'Success']):,} successful")
  if visit_corrections_df is not None and len(visit_corrections_df) > 0:
    print(f"Visit internal corrections applied: {len(visit_corrections_df[visit_corrections_df['Update_Status'] == 'Success']):,} successful")
  if visit_from_weed_corrections_df is not None and len(visit_from_weed_corrections_df) > 0:
    print(f"Visit-from-weed corrections applied: {len(visit_from_weed_corrections_df[visit_from_weed_corrections_df['Update_Status'] == 'Success']):,} successful")
  
  print("=" * 80 + "\n")
  
  return merged_df, overall_summary, field_summary, mismatches, corrections_df, visit_corrections_df, visit_from_weed_corrections_df


def main():
  parser = argparse.ArgumentParser(
    description="Analyze WeedLocations and Visits_Table field synchronization",
    formatter_class=argparse.RawDescriptionHelpFormatter,
    epilog="""
Examples:
  # Analyze and generate report only
  python weed_visits_analyzer.py --env development

  # Preview all corrections
  python weed_visits_analyzer.py --env development --correct-all --preview

  # Apply all corrections
  python weed_visits_analyzer.py --env development --correct-all

  # Apply specific correction types individually
  python weed_visits_analyzer.py --env development --correct-weed-from-visits
  python weed_visits_analyzer.py --env development --correct-visits
  python weed_visits_analyzer.py --env development --correct-visits-from-weed

  # Apply WeedLocations corrections to specific fields only
  python weed_visits_analyzer.py --env development --correct-weed-from-visits --fields "Urgency,Status"
    """
  )
  parser.add_argument(
    '--env',
    '--environment',
    dest='environment',
    required=False,
    help='Environment to analyze (e.g., development, production)'
  )
  parser.add_argument(
    '--output',
    '-o',
    dest='output_file',
    help='Output Excel file path (default: auto-generated with timestamp)'
  )
  parser.add_argument(
    '--ignore-dates',
    '--ignore-creation-edit-dates',
    dest='ignore_creation_edit_dates',
    action='store_true',
    help='Skip comparison of CreationDate_1 and EditDate_1 fields'
  )
  parser.add_argument(
    '--correct-weed-from-visits',
    '--correct-weed',
    dest='correct_mismatches',
    action='store_true',
    help='Update WeedLocations fields to match latest Visit data'
  )
  parser.add_argument(
    '--correct-visits',
    '--correct-visits-internal',
    dest='correct_visits',
    action='store_true',
    help='Fix Visits_Table internal data issues (e.g., set DateCheck from CreationDate_1 for all visits)'
  )
  parser.add_argument(
    '--correct-visits-from-weed',
    dest='correct_visits_from_weed',
    action='store_true',
    help='Update latest Visit fields from WeedLocations (e.g., copy ParentStatusWithDomain → WeedVisitStatus)'
  )
  parser.add_argument(
    '--correct-all',
    dest='correct_all',
    action='store_true',
    help='Apply all three correction types (equivalent to --correct-weed-from-visits --correct-visits --correct-visits-from-weed)'
  )
  parser.add_argument(
    '--preview',
    '--dry-run',
    dest='preview_only',
    action='store_true',
    help='Preview corrections without making changes (use with correction flags)'
  )
  parser.add_argument(
    '--fields',
    dest='fields_to_correct',
    help='Comma-separated list of WeedLocations field names to correct (use with --correct-weed-from-visits). Use --list-fields to see options.'
  )
  parser.add_argument(
    '--list-fields',
    action='store_true',
    help='List available field names and exit'
  )
  
  args = parser.parse_args()
  
  # Handle --list-fields (doesn't require environment)
  if args.list_fields:
    print("Available field names for --fields argument:")
    print("=" * 60)
    print(f"Latest Visit Strategy: {REFERENCE_DATE_STRATEGY}")
    print(f"  ({REFERENCE_DATE_DESCRIPTION})")
    print()
    for rule in FIELD_COMPARISON_RULES:
      category_tag = f" [{rule['category']}]" if rule.get('category') else ""
      print(f"  {rule['display_name']}{category_tag}")
      print(f"    WeedLocations: {rule['weed_field']}")
      print(f"    Visits_Table: {rule['visit_field']}")
      print(f"    Description: {rule['description']}")
      print()
    print("Note: Use --ignore-dates to skip [audit] fields")
    
    print("\nVisits_Table Internal Correction Rules (--correct-visits):")
    print("=" * 60)
    for rule in VISIT_CORRECTION_RULES:
      print(f"  {rule['display_name']}")
      print(f"    Target: {rule['target_field']}")
      print(f"    Source: {rule['source_field']}")
      print(f"    Description: {rule['description']}")
      print()
    
    print("Visits from WeedLocations Rules (--correct-visits-from-weed):")
    print("=" * 60)
    for rule in VISIT_FROM_WEED_RULES:
      print(f"  {rule['display_name']}")
      print(f"    Visit Field: {rule['visit_field']}")
      print(f"    Source (Weed): {rule['weed_field']}")
      print(f"    Description: {rule['description']}")
      print()
    return
  
  # Validate environment is provided for actual analysis/corrections
  if not args.environment:
    parser.error("--env is required (not needed for --list-fields)")
  
  # Handle --correct-all flag
  if args.correct_all:
    args.correct_mismatches = True
    args.correct_visits = True
    args.correct_visits_from_weed = True
  
  # Parse fields if provided
  fields_list = None
  if args.fields_to_correct:
    fields_list = [f.strip() for f in args.fields_to_correct.split(',')]
  
  # Validate flags
  if args.preview_only and not (args.correct_mismatches or args.correct_visits or args.correct_visits_from_weed):
    parser.error("--preview requires one of: --correct-weed-from-visits, --correct-visits, --correct-visits-from-weed, or --correct-all")
  
  if args.fields_to_correct and not args.correct_mismatches:
    parser.error("--fields requires --correct-weed-from-visits flag (only applies to WeedLocations corrections)")
  
  try:
    analyze_weed_visits(
      args.environment, 
      args.output_file, 
      args.ignore_creation_edit_dates,
      args.correct_mismatches,
      fields_list,
      args.correct_visits,
      args.correct_visits_from_weed,
      args.preview_only
    )
  except Exception as e:
    print(f"\nError during analysis: {e}")
    raise


if __name__ == "__main__":
  main()
